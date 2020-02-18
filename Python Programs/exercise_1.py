"""
mrp = 10
rating = 4.5
name = 'vikrant'
do_publish = True
star = '* ' * 10
print(mrp)
print(rating)
print(name)
print(do_publish)
print(star)

name = 'John Smith'
age = 25
new_patient = True
"""

# 2. Getting  Input ============================================


"""
name = input('What is you name : ')
color = input('What is your favourite color ? ')
print(name + ' likes ' + color)

birth_year = input('Please enter your birth year : ')
print(type(birth_year))
age = 2019 - int(birth_year) #taking input is always string
print(type(age))
print('You are ' + str(age) + ' years old') # only conca strings

wei_kg = input('Please enter you weigh in kg : ')
wei_po = float(wei_kg) * 2.20
print('Your weight in pound is : ' + str(round(wei_po, 2)))
"""

# 3. Strings ====================================================
"""
print("Python's course")
print('Python for "beginner"')
email = '''Hi John
This is our first email to you
Please check
'''
print(email)
course = 'Python course'
print(course[-4:]) # char from the last 4th
print(course[0:5]) # char starting to 5th
print(course[1:]) # char 1st to ending
print(course[:5]) # char starting to 5th
print(course[:]) # all characters
name = 'vikrant'
print(name[1:-1])
"""

# 4. Formatted Strings ====================================================
"""
fir_name = 'Vikrant'
las_name = 'Kumar'
mess = fir_name + ' [' + las_name + '] is a coder'
msg = f'{fir_name} [{las_name}] is a coder'
print(mess)
print(msg)
"""

# 5. String Methods ====================================================
"""
course = 'Python for beginners'
print(len(course))
print(course.upper())
print(course.lower())
print(course.find('n'))
print(course.find('beginners')) # to find a string
print(course.replace('beginners', 'absolute beginners')) # to replace a string
print(course.replace('n', 'N')) # to replace a char
print('Python' in course) # to find a string T/F
"""

# 6. Arithmetic Operators ==============================================

""""
a = 10
b = 3
print(a / b)
print(a // b) # modulus
print(a * b)
print(a ** b) # power
a += 3
print(a)
b = 10+3*2**2
c = (2 + 3) * 10 - 3**2
print(b)
print(c)
"""

# 6. Maths Functions ==================================================

"""
import math

a = -10.89908
print(abs(round(a,2)))
print(abs(math.ceil(a)))
print(math.floor(a))
"""
# 7. If statements ==================================================

"""
is_hot = True
is_cold = False
if is_hot :
    print("Its a hot day")
    print("Drink water")
elif is_cold:
    print("Its a cool day")
    print("Wear warm cloths")
else:
    print("its a lovely day")
print("Enjoy your day")


price = 1000000
has_good_credit = False

if has_good_credit :
    down = 0.1 * 1000000
else:
    down = 0.2 * 1000000
print(f'The down payment is ${round(down)}')
"""

# 8. Logical Operators ==================================================

"""
has_high_income = True
has_good_credit = False
has_criminal_record = False

if has_good_credit or has_high_income and not has_criminal_record :
    print("Eligible for loan")
else:
    print("Not eligible for loan")
"""

# 9. Comparision Operators ==================================================

"""
temp = 30
if temp >= 30 :
    print("Its a hot day")
elif temp <= 20 :
    print("Its a cold day")

name = input("Please enter you name : ")

if len(name) < 3 :
    print("Name must be at least 3 char")
elif len(name) > 50 :
    print("Name can be a maximum of 50 char")
else:
    print("Name looks good")
"""

# 10. Project Weight converter ==================================================

"""
weight = input("Please enter you weight : ")
val = input("(L)bs or (K)g : ")

if val.upper() == "L" :
    print("Your weight in KG : "+ str(round(float(weight) * 0.45, 2)))
elif val.upper() == "K" :
    print("Your weight in Pound : " + str(round(float(weight) / 0.45, 2)))
else :
    print("Invalid entry")
"""

# 11. While Loops ==================================================

"""
i = 1
while i <= 5 :
    print("*" * i)
    i = i + 1
print("done")

secret_number = 9
guess_count = 0
guess_limit = 3
while guess_count < guess_limit :
    guess = input("Guess : ")
    guess_count += 1
    if int(guess) == secret_number:
        print("You Won ! ")
        break
else:
    print("try next time")

command = ""
started = False
while True :
    command = input(">> ").lower()
    if command == "start" :
        if started :
            print("Already Started")
        else:
            started = True
            print("Car Started ")
    elif command == "stop" :
        if not started :
            print("Car already stopped")
        else:
            print("car stopped ")
            started = False
    elif command == "quit":
        break
    elif command == "help":
        print('''
start -> to start the car
stop  -> to stop the car
quit  -> to quit ''')
    else:
        print("I don't understand ")
"""

# 12. For Loops ==================================================

"""
for item in 'PYTHON' :
    print(item)

for item in ['vikrant', 'john', 'sarah'] :
    print(item)

for item in range(5,10, 2):
    print(item)

prices = [10, 20, 30, 40, 50]
total = 0
for item in prices :
    total += item
print(f"Total : {total}")

for x in range(4) :
    for y in range(3) :
        print(f"({x} , {y})")

num = [2, 2, 2, 2, 5]

for item in num:
    print('X'* item)
"""

# 13. List ==================================================

"""
num_list = [10, 22, 8, 5, 44, 16, 7, 25]
large_num = 0
i = 0
while i < len(num_list):
    if large_num < num_list[i] :
       large_num = num_list[i]
    i += 1
print(large_num)

num_list = [10, 22, 8, 5, 44, 16, 7, 25]
large_num = 0

for num in num_list :
    if large_num < num :
       large_num = num
print(large_num)
"""

# 14. 2D List ==================================================

"""
matrix = [

    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]
print(matrix[0][2])

for row in matrix :
    for item in row :
        print(item)
"""

# 15. List Methods ==================================================

"""
num_list = [10, 22, 8, 5, 44, 5, 7, 25]
num_list.sort()
num_list.reverse()
num_list.pop()
num_list.append(20)
num_list.insert(0, 17)
num_list.remove(5)
num_list.pop()
print(50 in num_list)
print(num_list.count(5))
print(num_list.)

num_list = [10, 22, 8, 5, 44, 5, 7, 25, 8]
unique = []

for num in num_list :
    if num not in unique :
        unique.append(num) 
print(unique)
"""

# 16. Tuples ================================================== similar ot list, can not edit

"""
num = (1, 2, 3)
print(num)
"""

# 17. Unpacking ==================================================

"""
cordinates = (1, 2, 3)
x, y, z = cordinates
"""

# 18. Dictionaries ==================================================

"""
customer = {
    "name": "vikrant",
    "age" : 30,
    "is_verfied" : True,
}
customer["name"] = "jack"
print(customer["name"])
print(customer.get("birth", "no bday"))

phone = input("Phone : ")
digit = {
    "1" : "One",
    "2" : "Two",
    "3" :  "three",
}

output = ""
for ch in phone :
    output = output + digit.get(ch, "") + " "
print(output)
"""

# 19. Emoji Converter ==================================================

"""
message = input("> ")
words = message.split(" ")
emoji = {
    ":)": "😊",
    ":(": "😒"
}
output = ""
for word in words:
    output += emoji.get(word, word) + " "
print(output)
"""

# 20. Functions  ==================================================

"""
def greet_user():
    print("Hi User")
    print("Welcome !!!")

print("start")
greet_user()
print("end")
"""

# 21.Parameters  ==================================================

"""
def greet_user(f_name,l_name):
    print(f'Hi {f_name} {l_name}')
    print('Welcome !!!')

print("start")
greet_user('John', 'smith')
print("end")
"""

# 22. Return Statement ==================================================

"""
def square(number):
    m = number*number
    return m

total = square(5)
print(total)
print(square(10))
"""

# 23. Exception ==================================================

"""
try :
    age = int(input('age = '))
    print(age)
    income = 20000
    risk = income / age
except ValueError :
    print("Invalid Value")
except ZeroDivisionError :
    print("Age cannot be 0")
"""

# 24. Classes ==================================================

"""
class Point:
    def move(self) :
        print("Move")
    def draw(self) :
        print("Draw")

point1 = Point()
point1.x = 10
print(point1.draw())
"""

# 25. Constructor ==================================================

"""
class Point:
    def __init__(self, x, y):
        self.x = x
        self.y = y
    def draw(self):
        print(f'Hi I am {self.x} {self.y}')
point1 = Point('vikrant', 'kumar')
point1.draw()
point2 = Point('Bob' , 'ray')
point2.draw()
"""

# 26. Inheritance ==================================================

"""
class Mammal:
    def walk(self):
        print("walk")

class Dog(Mammal):
   def bark(self):
       print("barks")

class Cat(Mammal):
    def mem(self):
        print("be annoying")

dog1 = Dog()
dog1.bark()
dog1.walk()
"""

# 26. Modules ==================================================

"""
import utils
from utils import lbs_to_kg

print(utils.kg_to_lbs(45))
print(lbs_to_kg(100))

from utils import find_max

max_num = find_max([10, 5, 3, 6, 12])
print(max_num)
"""

# 27. Packages ==================================================

"""
import ecommerce.shipping
from ecommerce import shipping

shipping.calc_shipping()
"""

# 28. Modules ==================================================
# PYTHON 3 MODULE INDEX

"""
import random

members = ['Vikrant', 'Riya', 'vikriy']
for i in range(3):
    print(random.randint(10, 20))
    print((random.choice(members)))

import random
class Dice:
    def roll(self):
        first = random.randint(1,6)
        second = random.randint(1,6)
        return first, second

roller = Dice()
print(roller.roll())
"""

# 29. Files and Directories ==================================================

"""
from pathlib import Path

# Absolute Path : the complete path
# Relative path : Starting from current dir

path = Path("ecommerce")
for file in (path.glob('*')):
    print(file)
"""

# 30. Pypi and Pip(Huh?) ==================================================
#pypi.org
# to get the pre configured programs

# 31. Excel Project ==================================================

"""
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    cell = sheet['A1']
    #print(sheet.max_row)
    for row in range(2 , sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value*0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save('trans.xlsx')
process('transaction.xlsx')
"""

# 32. What is machine learning ==================================================
# 32. Machine learning in action ================================================
# 33. Library and tools =========================================================
    #1. Numpy        : Provides a multi dimensional array.
    #2. Pandas       : Data Analysis like data frame
    #3. MatPlotLib   : 2 dimensional plotting library , helps to create graphs and plots
    #4. Scikit-Learn : helps for algorithms like decision tree, neural network

    #1. Jupyter : IDe for machine learning
    #2. Anaconda
    #3. Kaggle : A dataset website
        
    #4. panda dataframe

#34. Import a data set =========================================================














